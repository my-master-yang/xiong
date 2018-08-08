__authon__ = 'geyu'
__date__ = '18-6-7 '

from .models import Question, ExamCategory
from openpyxl import Workbook, load_workbook
# from openpyxl.utils import get_column_letter
from openpyxl.compat import range
import random


# 导入excel文件　批量存入数据库
def import_examFile(self, request, obj, change):
    wb = load_workbook(filename=obj.examFile.path)
    ws = wb.get_sheet_names()
    # print(ws)
    ws = wb.get_sheet_by_name(ws[0])
    ws_rows_len = ws.max_row
    headers = [
        'ExamCategory', 'ExamPaperId', 'questionType', 'curFraction', 'questionContent', 'answer', 'optionA',
        'optionB', 'optionC', 'optionD'
    ]
    lists = []
    # users = request.user
    for row in range(2, ws_rows_len + 1):
        r = {}
        for col in range(1, len(headers) + 1):
            key = headers[col - 1]
            r[key] = ws.cell(row=row, column=col).value
        lists.append(r)
    sqllist = []
    for cell in lists:
        # for header in headers:
        examcategory = ExamCategory.objects.get(name=cell['ExamCategory'])
        exampaperId = cell['ExamPaperId']
        # questionId = cell['QuestionId']
        questionType = cell['questionType']
        curFraction = cell['curFraction']
        questionContent = cell['questionContent']
        answer = cell['answer']
        optionA = cell['optionA']
        optionB = cell['optionB']
        optionC = cell['optionC']
        optionD = cell['optionD']

        sql = Question(
            examcategory=examcategory,
            examLessonNum=exampaperId,
            questionType=questionType,
            score=curFraction,
            content=questionContent,
            answer=answer,
            OptionA=optionA,
            OptionB=optionB,
            OptionC=optionC,
            OptionD=optionD)
        if (Question.objects.filter(content=questionContent).first() != None):
            continue
        else:
            sqllist.append(sql)
    Question.objects.bulk_create(sqllist)


# 从考试题库随机生成试卷的测试题
def ExamRandomPaperView(self, request):

    questionss = Question.objects.all().values('id').filter(examcategory_id=8)

    qu_list = list(questionss)  # 将QuerySet转换成list
    randomquestionss = []
    for i in range(0, 30):  # 随机抽取30道题
        t = random.choice(qu_list)  # 从list中随机选取一个元素
        randomquestionss.append(t)
        qu_list.remove(t)  # 将已抽取的元素删除,确保不重复抽取

    # print("\nRAN-CHOICEQUE:", randomquestionss, " RQCOUNT:", len(randomquestionss))

    return randomquestionss
